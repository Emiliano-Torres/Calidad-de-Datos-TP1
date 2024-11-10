import pandas as pd
from openpyxl import load_workbook

# Cargar el archivo de Excel con pandas
archivo = 'GSAF5.xlsx'
GSAF= pd.read_excel(archivo, sheet_name='Sheet1-GSAF')  

# Cargar el archivo de Excel con openpyxl
libro = load_workbook(archivo, data_only=True)
hoja = libro['Sheet1-GSAF'] 

# Función para obtener el color de fondo de una celda específica
def obtener_color_fondo(fila, columna):
    celda = hoja.cell(row=fila + 1, column=columna + 1)  # Las celdas en openpyxl empiezan desde 1
    return celda.fill.start_color.rgb if celda.fill.start_color else None

# Crear una nueva columna en el DataFrame con los colores de cada celda en la primera columna
GSAF=GSAF[['Date','year','State','Location','Type','Activity','Name','Sex','Age','Injury','Fatal','Time','Species','Source']]
GSAF['ColorFondo'] = [obtener_color_fondo(i, 0) for i in range(len(GSAF))]
GSAF.to_excel('GSAF5.xlsx',index =False)